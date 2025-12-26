from typing import Dict
from openpyxl.worksheet.worksheet import Worksheet


def parse_down(sheet: Worksheet, i, j, r, _, ftype):
    course_map = {}
    if ftype == 1:
        while i <= r:
            v = sheet.cell(i, j).value
            v2 = sheet.cell(i, j + 1).value
            s = sheet.cell(i, j + 2).value
            i += 1

            if v2 == "(19M21HS111)":
                course_map.update({str(v2).strip("() "): str(s)})

            if v is None or v2 is None or s is None:
                continue

            v = str(v).strip("() ")
            v2 = str(v2).strip("() ")
            s = str(s).strip()
            update(course_map, v, s, v2, s)
            # course_map.update({v: s, v2: s})
    elif ftype == 2:
        while i <= r:
            v = str(sheet.cell(i, j).value)
            s = str(sheet.cell(i, j + 1).value).strip(" -")
            if v == "Faculty Abbreviation with Names" or v == "Faculty Abbreviation":
                break

            i += 1
            if v is None or s is None:
                continue

            v = v.replace(" ", "")
            v = v.replace("\n", "")
            if "/" in v:
                v1, v2 = v.split("/")
            else:
                v1 = v2 = v
            
            update(course_map, v1.strip(), s, v2.strip(), s, v.strip(), s)
            # course_map.update({v1.strip(): s, v2.strip(): s, v.strip(): s})
    elif ftype == 3:
        while i <= r:
            v = str(sheet.cell(i, j).value).strip().replace(" ", "").replace("\n", "")
            s = str(sheet.cell(i, j + 1).value).strip()
            if not (len(v) > 3 and v[:2].isnumeric() and v[2:3].isalpha()):
                break
            i += 1

            v1 = v
            v2 = v[2:]
            v3 = v[5:]

            update(course_map, v, s, v2, s, v3, s)
            course_map.update({v: s, v2: s, v3: s})

    elif ftype == 4:
        while i <= r:
            code = sheet.cell(i, j).value
            name = sheet.cell(i, j - 1).value
            i += 1
            if code is None or name is None:
                continue
            update(course_map, str(code).strip(), str(name).strip())
            # course_map.update({str(code).strip(): str(name).strip()})

    elif ftype == 5:
        while i <= r:
            short = sheet.cell(i, j).value
            name = sheet.cell(i, j + 2).value
            i += 1
            if not short or not name:
                continue

            short = str(short).strip()
            name = str(name).strip()
            update(course_map, short, name)
            # course_map[short] = name

    return course_map


def parse_courses(sheet: Worksheet, row: int, col: int):
    course_map = {}
    for i in range(1, row + 1):
        for j in range(1, col + 1):
            pvalue = None
            if j - 1 > 0:
                pvalue = str(sheet.cell(i, j - 1).value)
            value = str(sheet.cell(i, j).value)
            nvalue = str(sheet.cell(i, j + 1).value)
            if value in ["Short Subject Code", "CODE", "SHORT FORM"] and (
                nvalue in ["Subject Code", "SUBJECT CODE"]
            ):
                course_map.update(parse_down(sheet, i + 1, j, row, col, ftype=1))
            elif value in ["SHORT FORM / SUBJECT CODE", "SHORT FORM /"] and nvalue in ["SUBJECT NAME"]:
                course_map.update(parse_down(sheet, i + 1, j, row, col, ftype=2))
            elif (
                pvalue in ["Name"]
                and value in ["SUBJECT CODE"]
                and nvalue in ["SUBJECT NAME"]
            ):
                course_map.update(parse_down(sheet, i + 1, j, row, col, ftype=2))
            elif (
                pvalue not in ["Name"]
                and value in ["SUBJECT CODE"]
                and nvalue in ["SUBJECT NAME"]
            ):
                course_map.update(parse_down(sheet, i + 1, j, row, col, ftype=3))

            elif pvalue in ["COURSE"] and value in ["COURSE CODE"]:
                course_map.update(parse_down(sheet, i + 1, j, row, col, ftype=4))

            elif value == "Short Name" and nvalue == "Course Code":
                course_map.update(parse_down(sheet, i + 1, j, row, col, ftype=1))
    return course_map


def update(map: Dict[str, str], *keyvalues: str):
    if len(keyvalues) % 2 != 0:
        return

    for i in range(0, len(keyvalues), 2):
        key = keyvalues[i]
        value = keyvalues[i + 1]
        if len(key) == 10: # if full sized code then add its derivatives
            map.update({key: value, key[2:]: value, key[3:]: value})
        else:
            map[key] = value

