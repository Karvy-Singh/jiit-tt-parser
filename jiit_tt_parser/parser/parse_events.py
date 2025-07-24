import datetime
from typing import Dict, Literal, List
import string
import re

from openpyxl.cell import Cell, MergedCell
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from jiit_tt_parser.parser.parse_courses import parse_courses
from jiit_tt_parser.parser.parse_electives import parse_electives
from jiit_tt_parser.utils.utils import are_cells_in_same_merged_group, is_empty_row
from jiit_tt_parser.utils.cache import load_map, FACULTY_MAP

days_of_the_week_names = [
    "monday",
    "tuesday",
    "wednesday",
    "thursday",
    "friday",
    "saturday",
    "sunday",
]


class Period:
    def __init__(
        self, start: datetime.time | None = None, end: datetime.time | None = None
    ) -> None:
        self.start_time = start or datetime.time(0, 0)
        self.end_time = end or datetime.time(23, 59)

    @classmethod
    def from_string(cls, fmt: str):
        start, end = fmt.split("-")
        start = start.strip(" NO")
        end = end.strip("APM ")

        end_hour, end_min = split_hour_min(end)
        start_hour, start_min = split_hour_min(start)

        if start_hour < 8:
            start_hour += 12
        if end_hour < 9:
            end_hour += 12
        start_time = datetime.time(start_hour, start_min)
        end_time = datetime.time(end_hour, end_min)

        return cls(start_time, end_time)

    def __add__(self, other):
        start_time = other.start_time
        end_time = other.end_time
        if self.start_time < other.start_time:
            start_time = self.start_time

        if self.end_time > other.end_time:
            end_time = self.end_time

        return Period(start_time, end_time)

    def __str__(self):
        return f"{self.start_time.hour}:{str(self.start_time.minute).zfill(2)} - {self.end_time.hour}:{str(self.end_time.minute).zfill(2)}"


class Event:
    def __init__(self, event_string: str):
        self.event_string = event_string
        self.batches: List[str]
        self.event_type: Literal["L", "T", "P", "TALK"] | str  # just to please linter
        self.classroom: str
        self.event: str
        self.eventcode: str
        self.period: Period
        self.day: str
        self.lecturer: List[str]

    @classmethod
    def from_string(
        cls,
        ev_str: str,
        period: Period,
        day: str,
        courses: dict,
        _: str,  # elective_category
        faculties: dict,
    ):
        ev_str = ev_str.strip().replace("\n", " ").replace("\xa0", " ")
        print(repr(ev_str))
        og = ev_str

        if "C1-C3HS" in ev_str:
            ev_str = ev_str.replace("C1-C3HS", "C1-C3(HS")

        if "LC1-C3(HS211)-/FF1KMB" in ev_str:
            ev_str = "LC1-C3(HS211)-/FF1/KMB"

        if ev_str.startswith("PBG") and ev_str[3].isdigit():
            ev_str = "PG" + ev_str[3:]

        if "A5-A6-A10" in ev_str:
            ev_str = ev_str.replace("A5-A6-A10", "A5,A6,A10")

        if ev_str == "":
            return None
        ev = cls(ev_str)
        if "TALK" in ev_str:
            ev.event_type = "TALK"
            ev.batches, ev_str = (
                parse_batches(ev_str[: ev_str.find("(")]),
                ev_str[ev_str.find("(") :],
            )
            ev.batches = [i.strip() for i in ev.batches]
            ev.eventcode = "TALK"
            ev.event = "Talk"
            ev.classroom = ev_str.split("-")[-1].strip()
            ev.lecturer = []

            ev.period = period
            ev.day = day.capitalize()
            return ev

        ev.event_type, ev_str = ev_str[:1], ev_str[1:]

        raw_batches = ev_str[: ev_str.find("(")]
        ev_str = ev_str[ev_str.find("(") :]

        if "." in raw_batches:
            raw_batches = raw_batches.replace(".", ",")
        ev.batches = parse_batches(raw_batches)

        ev.batches = [i.strip() for i in ev.batches]

        # if len(ev.batches) > 1:
        #     for i, b in enumerate(ev.batches[1:], start=1):
        #         if b[0].isdigit():
        #             ev.batches[i] = f"{ev.batches[0][0]}{ev.batches[i]}"
        ev.eventcode, ev_str = ev_str[1 : ev_str.find(")")], ev_str[ev_str.find(")") :]
        ev.eventcode = ev.eventcode.strip()
        if ev.eventcode == "M302":
            ev.eventcode = "MA302"
        ev.event = lookup_sub(ev.eventcode.strip(), courses) or ""
        ev.event = " ".join(
            ev.event.replace("\xa0", " ").replace("\n", " ").strip().split()
        )

        while ev_str[0].upper() not in string.ascii_uppercase + string.digits:
            ev_str = ev_str[1:]

        classroom: str
        teacher: str
        ok: bool
        classroom, teacher, ok = parse_classroom_teacher_format(ev_str)
        if ok:
            subs = [classroom, teacher]
        else:
            subs = extract_substrings(ev_str)

        if any(
            substr in ev_str
            for substr in ["EDD/CADD0", "ACL,JBSPL", "SPL, 5G LAB/", "SPL,5G LAB/"]
        ):  # case of two classrooms
            ev.classroom = "/".join(subs[:2])
            subs = subs[2:]
        elif any(
            substr in ev_str for substr in ["BS,SHG/CL15/CL16"]
        ):  # case of reversed two classrooms
            ev.classroom = "/".join(subs[-2:])
            subs = subs[:-2]
        elif og == "TA18(25B31EC311)-TA13/MO":
            ev.classroom = "N/A"
        elif "PL2/RAV.NFP1" in ev_str:
            ev.classroom = subs[0]
            subs = subs[1].split(".")
        elif "SR05 NFMATHS3" in ev_str:
            a = subs[0].split()
            ev.classroom = a[0]
            subs = [a[1]]
        else:
            ev.classroom, subs = parse_class_and_faculty(subs)

        ev.lecturer = subs
        for i in range(len(ev.lecturer)):
            lecturer = ev.lecturer[i].strip("- ")
            v = faculties.get(lecturer)
            if v is not None:
                lecturer = v.title()

            ev.lecturer[i] = " ".join(
                lecturer.replace("\xa0", " ").replace("\n", " ").strip("- ").split()
            )
            nf, ok = get_new_faculty(ev.lecturer[i])
            if ok:
                ev.lecturer[i] = nf

            ta, ok = get_teaching_assistant(ev.lecturer[i])
            if ok:
                ev.lecturer[i] = ta

        ev.period = period
        ev.day = day.capitalize()
        if ev.event == "":
            print(repr(og))
        print(ev)

        return ev

    def __str__(self) -> str:
        # print(repr(self.event_type))
        # print(self.event_string)
        lecture_types = {
            "L": "Lecture",
            "T": "Tutorial",
            "P": "Practical",
            "TALK": "Talk",
        }

        return f"""Event: {lecture_types[self.event_type]}
Time: {self.period}
Day: {self.day}
Batches: {self.batches}
Subject: {self.event or self.eventcode}
Venue: {self.classroom}
Lecturer: {self.lecturer}
"""


def extract_substrings(input_string, delimiters=",/\\"):
    """
    Extract substrings from a string separated by specified delimiters.

    Args:
        input_string (str): String with substrings separated by delimiters
        delimiters (str): String containing delimiter characters (default: ',/\\-')

    Returns:
        list: List of extracted substrings
    """
    if not input_string:
        return []

    if not delimiters:
        return [input_string.strip()] if input_string.strip() else []

    # Escape special regex characters in delimiters
    escaped_delimiters = re.escape(delimiters)

    # Create regex pattern: [escaped_delimiters]+
    pattern = f"[{escaped_delimiters}]+"

    # Split by any combination of the specified delimiters
    substrings = re.split(pattern, input_string)

    # Filter out empty strings that might result from multiple consecutive delimiters
    result = [substring.strip() for substring in substrings if substring.strip()]

    return result


def parse_batches(batch_string):
    """
    Parse a batch string and return a list of individual batch codes.

    Args:
        batch_string (str): String containing batch specifications separated by commas

    Returns:
        list: List of individual batch codes

    Raises:
        ValueError: If range has mismatched letters or invalid format
    """
    if not batch_string or not batch_string.strip():
        return []

    # Remove all spaces
    batch_string = batch_string.replace(" ", "")

    # Split by commas
    parts = [part.strip() for part in batch_string.split(",") if part.strip()]

    # Check if all parts are string-only format (no letter+number pattern)
    if all(is_string_only_format(part) for part in parts):
        return parts

    # Otherwise, process as regular format
    result = []
    current_letter = None

    for part in parts:
        if not part:
            continue

        # Check if it's a range (contains hyphen)
        if "-" in part:
            result.extend(parse_range(part))
            # Update current letter from the range
            range_match = re.match(r"^([A-Z])(\d+)-([A-Z]?)(\d+)$", part)
            if range_match:
                current_letter = range_match.group(1)

        # Check if it's a concatenated format like F2F3F6F7
        elif re.match(r"^[A-Z]\d+([A-Z]\d+)+$", part):
            result.extend(parse_concatenated(part))
            # Update current letter from the last batch in concatenated format
            matches = re.findall(r"([A-Z])(\d+)", part)
            if matches:
                current_letter = matches[-1][0]

        # Check if it's just a number (inherits current letter)
        elif part.isdigit():
            if current_letter is None:
                raise ValueError(f"Number '{part}' found without a preceding letter")
            result.append(f"{current_letter}{part}")

        # Check if it's a single batch code
        elif re.match(r"^[A-Z]\d+$", part):
            result.append(part)
            # Update current letter
            current_letter = part[0]

        else:
            raise ValueError(f"Invalid batch format: '{part}'")

    return result


def is_string_only_format(part):
    """
    Check if a part is in string-only format (no letter+number pattern).
    String-only format means it doesn't match the regular batch pattern.
    """
    # If it contains only letters, hyphens, and possibly other non-digit characters
    # and doesn't match the standard batch pattern, consider it string-only
    return (
        not re.match(r"^[A-Z]\d+$", part)
        and not re.search(r"[A-Z]\d+", part)
        and not part.isdigit()
    )


def parse_range(range_str):
    """Parse a range like 'C1-C3' or 'C1-3' into individual batches."""
    # Match patterns like C1-C3 or C1-3
    match = re.match(r"^([A-Z])(\d+)-([A-Z]?)(\d+)$", range_str)

    if not match:
        raise ValueError(f"Invalid range format: '{range_str}'")

    start_letter = match.group(1)
    start_num = int(match.group(2))
    end_letter = match.group(3)
    end_num = int(match.group(4))

    # If end letter is provided, it must match start letter
    if end_letter and end_letter != start_letter:
        raise ValueError(f"Range cannot span different letters: '{range_str}'")

    # Validate range order
    if start_num > end_num:
        raise ValueError(f"Invalid range order: '{range_str}'")

    # Generate the range
    return [f"{start_letter}{i}" for i in range(start_num, end_num + 1)]


def parse_concatenated(concat_str):
    """Parse concatenated format like 'F2F3F6F7' into individual batches."""
    # Find all batch codes in the concatenated string
    matches = re.findall(r"([A-Z])(\d+)", concat_str)

    if not matches:
        raise ValueError(f"No valid batch codes found in: '{concat_str}'")

    return [f"{letter}{number}" for letter, number in matches]


def get_time_row(sheet: Worksheet, row, col):
    for i in range(1, row + 1):
        v = sheet.cell(i, 2).value
        if str(v).startswith("9") or str(v).startswith("8"):
            for j in range(2, col + 1):
                if sheet.cell(i, j).value is None:
                    return i, j - 1
            return i, col
    return 2, col


def get_day_row(sheet: Worksheet, row, _, day: str):
    day = day.lower()
    for i in range(1, row + 1):
        v = str(sheet.cell(i, 1).value).lower()
        if day.startswith(v):
            return i

    return -1


def get_periods(sheet: Worksheet, _, col, time_row):
    a = []
    for i in range(2, col + 1):
        p = Period.from_string(str(sheet.cell(time_row, i).value))
        a.append(p)

    return a


def is_end_of_day(sheet: Worksheet, curr, day, cols):
    # if curr >= 300:
    #     return True

    if day.lower() != "saturday":
        return sheet.cell(curr + 1, 1).value is not None

    # v = sheet.cell(curr, 1).value
    # print(v)
    theme = sheet.cell(curr, 1).fill.start_color.theme
    if theme is not None and theme == 1:
        return True

    if isinstance(sheet.cell(curr, 1), MergedCell):
        if not are_cells_in_same_merged_group(
            sheet, sheet.cell(curr, 1), sheet.cell(curr + 1, 1)
        ):
            return True
    else:
        if is_empty_row(sheet, curr, cols):
            return True

    return False


def search_merged_cells(merged_cells: list[CellRange], cell: Cell) -> int | None:
    for c in merged_cells:
        if c.min_row != c.max_row:
            continue

        if (c.min_row == cell.row) and (
            cell.column >= c.min_col and cell.column <= c.max_col
        ):
            return c.max_col

    return None


def parse_day(
    sheet: Worksheet,
    _: int,  # row
    col: int,
    start,
    periods: List[Period],
    day: str,
    merged_cells: list[CellRange],
    courses: dict,
    faculties: dict,
):
    spam_entries = [
        "LUNCH",
        "ALL BATCH FREE FOR MEETING",
        "FREE TS11",
        "/NFMATH3",
        "BLOCKED",
        "LECTURE AND TUTORIAL CLASSES ARE BLOCKED FOR TALKS.",
    ]

    elective_categories = [
        "SE",
        "HSS 1",
        "HSS1",
        "HSS-1",
        "HSS 2",
        "HSS2",
        "HSS-2",
        "OE 2",
        "OE2",
        "OE-2",
        "DE 1",
        "DE1",
        "DE-1",
        "DE 2",
        "DE2",
        "DE-2",
        "DE 3",
        "DE3",
        "DE-3",
        "DE 4",
        "DE4",
        "DE-4",
        "DE 5",
        "DE5",
        "DE-5",
        "DE 6",
        "DE6",
        "DE-6",
    ]
    events = []
    if str(sheet.cell(start, 2).value).startswith("9"):
        start += 1

    for j in range(2, col + 1):
        r = start
        # elective_cat = ""
        reached_end = False
        while not reached_end:
            reached_end = is_end_of_day(sheet, r, day, col)
            c = sheet.cell(r, j)
            r += 1
            if (v := c.value) is None:
                continue
            ev_str = str(v).replace("\xa0", " ").replace("\n", " ").strip().upper()

            if ev_str in elective_categories:
                break

            if ev_str in spam_entries:
                continue

            if not any(ch.isalpha() for ch in ev_str):
                continue

            ep = periods[j - 2]
            if m := search_merged_cells(merged_cells, c):
                ep += periods[m - 2]

            ev = Event.from_string(ev_str, ep, day, courses, "", faculties)
            if ev is None:
                continue
            events.append(ev)

    return events


def parse_events(
    sheet: Worksheet,
    electives_file: str,
    row: int,
    col: int,
    faculty_map_path: str = FACULTY_MAP,
    curriculum_map_path: str = "curriculum.json",
) -> List[Event]:
    time_row, col = get_time_row(sheet, row, col)
    periods = get_periods(sheet, row, col, time_row)
    merged_cells = sheet.merged_cells.sorted()
    courses = parse_courses(sheet, row, col)
    _ = parse_electives(electives_file)  # electives
    faculties = load_map(faculty_map_path)
    curriculum = load_map(curriculum_map_path)
    curriculum_courses: Dict[str, str] = curriculum["courses"]
    curriculum_courses.update(courses)
    curriculum_courses["EC112"] = "Basic Electronics for Biotechnology"

    events = []

    for day in days_of_the_week_names:
        r = get_day_row(sheet, row, col, day)
        if r < 0:
            continue

        events.extend(
            parse_day(
                sheet,
                row,
                col,
                r,
                periods,
                day,
                merged_cells,
                curriculum_courses,
                faculties,
            )
        )

    return events


def split_hour_min(time_str):
    if ":" in time_str or "." in time_str:
        parts = re.split(r"[:.]", time_str)
        hour = int(parts[0])
        minute = int(parts[1].strip("AMP ")) if len(parts) > 1 else 0
    else:
        hour = int(time_str)
        minute = 0
    return hour, minute


def contains_number(s):
    return any(char.isdigit() for char in s)


def get_teaching_assistant(input_string: str) -> tuple[str, bool]:
    """
    Check if string matches format 'TA{number}' or 'TA-{number}' and extract components.

    Args:
        input_string (str): String to check

    Returns:
        tuple: (formatted_string, is_valid)
               - formatted_string: "Teaching Assistant {number}" if valid, anything if invalid
               - is_valid: True if format matches, False otherwise
    """
    if not input_string:
        return "Invalid input", False

    # Pattern to match TA followed by optional hyphen and then digits
    pattern = r"^TA-?(\d+)$"
    match = re.match(pattern, input_string)

    if match:
        number = match.group(1)
        return f"Teaching Assistant {number}", True
    else:
        return "Invalid format", False


def get_new_faculty(input_string) -> tuple[str, bool]:
    """Check if string matches format 'NF{chars}{number}' and extract components."""
    if not input_string:
        return "Invalid input", False

    pattern = r"^NF(.*?)(\d+)$"
    match = re.match(pattern, input_string)

    char_map = {"P": "Physics", "M": "Maths"}

    if match:
        chars = match.group(1)
        number = match.group(2)

        if chars:
            if char_map.get(chars) is not None:
                return f"New Faculty {char_map[chars]} {number}", True
            return f"New Faculty {chars} {number}", True
        else:
            return f"New Faculty {number}", True
    else:
        return "Invalid format", False


def is_classroom_code(text):
    """Check if text matches classroom code pattern {2+ chars}{number}"""
    return bool(re.match(r"^[A-Z]{2,}\d+$", text))


def is_standalone_number(text):
    """Check if text is just a number"""
    return text.isdigit()


def parse_class_and_faculty(string_list: List[str]) -> tuple[str, List[str]]:
    """
    Parse a list of strings to determine class code and faculty list.
    Handles classroom concatenation like CL10,11 -> CL10/CL11

    Args:
        string_list (list): List of strings containing class code and faculty

    Returns:
        tuple: (class_code, faculty_list)
               - class_code: The identified class code (may be concatenated classrooms)
               - faculty_list: List of faculty members
    """
    if not string_list or len(string_list) == 0:
        return "", []

    if len(string_list) == 1:
        # Only one element, assume it's class code
        return string_list[0], []

    # Check for classroom concatenation pattern at the end
    # Look for: [...faculty, classroom_code, number1, number2, ...]
    classroom_concat_result = check_classroom_concatenation(string_list)
    if classroom_concat_result:
        return classroom_concat_result

    first_element = string_list[0]
    last_element = string_list[-1]

    # Case 1: Check if first element is New Faculty or Teaching Assistant
    _, is_new_faculty = get_new_faculty(first_element)
    _, is_teaching_assistant = get_teaching_assistant(first_element)

    if is_new_faculty or is_teaching_assistant:
        # First element is faculty, so last element is class code
        return last_element, string_list[:-1]

    # Case 2: Check if last element matches {2 chars}{number} format (but not NF or TA)
    last_element_pattern = r"^([A-Z]{2})(\d+)$"
    match = re.match(last_element_pattern, last_element)

    if match:
        prefix = match.group(1)
        # Check if prefix is NOT "NF" or "TA"
        if prefix not in ["NF", "TA"]:
            # Last element is class code, rest are faculty
            return last_element, string_list[:-1]

    # Default case: First element is class code, rest are faculty
    return first_element, string_list[1:]


def check_classroom_concatenation(
    string_list: List[str],
) -> tuple[str, List[str]] | None:
    """
    Check for classroom concatenation pattern and handle it.
    Pattern: [...faculty, classroom_code, number1, number2, ...]
    Returns: (concatenated_classroom, faculty_list) or None if pattern not found
    """
    if len(string_list) < 3:
        return None

    # Look for a classroom code followed by one or more standalone numbers
    for i in range(
        1, len(string_list) - 1
    ):  # Start from index 1, don't check last element alone
        current = string_list[i]

        # Check if current element is a classroom code
        if is_classroom_code(current):
            # Check if all remaining elements are standalone numbers
            remaining_elements = string_list[i + 1 :]
            if all(is_standalone_number(elem) for elem in remaining_elements):
                # Extract the base from classroom code
                match = re.match(r"^([A-Z]+)(\d+)$", current)
                if match:
                    base_letters = match.group(1)
                    # base_number = match.group(2)

                    # Build concatenated classroom string
                    classrooms = [current]  # Start with the full classroom code
                    for num in remaining_elements:
                        classrooms.append(f"{base_letters}{num}")

                    concatenated_classroom = "/".join(classrooms)
                    faculty_list = string_list[
                        :i
                    ]  # Everything before the classroom code

                    return concatenated_classroom, faculty_list

    return None


def parse_classroom_teacher_format(input_string: str) -> tuple[str, str, bool]:
    """
    Check if string matches format '{any chars}{any number}{whatever follows starting with letter}'
    and extract classroom and teacher components.

    Args:
        input_string (str): String to check and parse

    Returns:
        tuple: (classroom, teacher, is_valid)
               - classroom: The chars + number part if valid
               - teacher: The remaining part if valid
               - is_valid: True if format matches, False otherwise
    """
    if not input_string:
        return "", "", False

    # Pattern: any letters followed by any digits followed by anything starting with a letter
    pattern = r"^([A-Za-z]+)(\d+)([A-Za-z].*)$"
    match = re.match(pattern, input_string)

    if match:
        chars = match.group(1)
        number = match.group(2)
        teacher = match.group(3)

        classroom = chars + number
        return classroom, teacher, True
    else:
        return "", "", False


def check_list_classroom_teacher_format(string_list):
    """
    Check if all strings in the list follow the classroom-teacher format.

    Args:
        string_list (list): List of strings to check

    Returns:
        tuple: (parsed_list, all_match)
               - parsed_list: List of (classroom, teacher) tuples
               - all_match: True if all strings match the format
    """
    if not string_list:
        return [], False

    parsed_list = []
    all_match = True

    for s in string_list:
        classroom, teacher, is_valid = parse_classroom_teacher_format(s)
        if is_valid:
            parsed_list.append((classroom, teacher))
        else:
            all_match = False
            break

    return parsed_list, all_match


def lookup_sub(subject_code, subject_dict):
    """
    Lookup subject name from potentially malformed subject code.

    Args:
        subject_code (str): The subject code to lookup
        subject_dict (dict): Dictionary mapping subject codes to subject names

    Returns:
        str or None: Subject name if found, None if not found
    """
    if not subject_code or not subject_dict:
        return None

    if (
        v := subject_dict.get(subject_code)
    ) is not None:  # first do a simple lookup, if its works good
        return v
    # Clean the input
    subject_code = subject_code.strip()

    # Classify the format
    format_type = classify_format(subject_code)

    if format_type == 1:  # Full format: 15B11CI111
        # Fix malformed code and try lookups
        return lookup_full_format(subject_code, subject_dict)
    elif format_type == 2:  # Medium format: B11CI111
        return lookup_medium_format(subject_code, subject_dict)
    elif format_type == 3:  # Short format: CI111
        return lookup_short_format(subject_code, subject_dict)
    else:
        # Unknown format, try direct lookup
        return subject_dict.get(subject_code)


def classify_format(code):
    """Classify the format of the subject code."""
    # Format 1: 15B11CI111 (2 digits + char + 2 digits + 2 chars + 3-4 digits)
    if re.match(r"^\d{2}[A-Z]\d{1,2}[A-Z]{2}\d{3,4}$", code):
        return 1

    # Format 2: B11CI111 (char + 2 digits + 2 chars + 3-4 digits)
    if re.match(r"^[A-Z]\d{1,2}[A-Z]{2}\d{3,4}$", code):
        return 2

    # Format 3: CI111 (2 chars + 3-4 digits)
    if re.match(r"^[A-Z]{2}\d{3,4}$", code):
        return 3

    return 0  # Unknown format


def fix_malformed_code(code):
    """Fix malformed full format codes and return list of possible corrections."""
    # Extract parts using regex
    match = re.match(r"^(\d{2})([A-Z])(\d{1,2})([A-Z]{2})(\d{3,4})$", code)
    if not match:
        return [code]

    prefix = match.group(1)  # 2 digits
    middle_char = match.group(2)  # single char
    middle_digits = match.group(3)  # 1-2 digits
    chars = match.group(4)  # 2 chars
    end_digits = match.group(5)  # 3-4 digits

    # Fix end digits if 4 digits (truncate to 3)
    if len(end_digits) == 4:
        end_digits = end_digits[:3]

    possible_codes = []

    # If middle_digits is only 1 digit, try adding digits
    if len(middle_digits) == 1:
        original_digit = middle_digits

        # Try adding after the digit (0-9)
        for i in range(10):
            new_middle = original_digit + str(i)
            fixed_code = f"{prefix}{middle_char}{new_middle}{chars}{end_digits}"
            possible_codes.append(fixed_code)

        # Try adding before the digit (0-9)
        for i in range(10):
            new_middle = str(i) + original_digit
            fixed_code = f"{prefix}{middle_char}{new_middle}{chars}{end_digits}"
            possible_codes.append(fixed_code)
    else:
        # Middle digits are already 2, just fix end digits if needed
        fixed_code = f"{prefix}{middle_char}{middle_digits}{chars}{end_digits}"
        possible_codes.append(fixed_code)

    return possible_codes


def lookup_full_format(code, subject_dict):
    """Lookup full format with malformation fixing and fallbacks."""
    # Get all possible fixed versions
    possible_codes = fix_malformed_code(code)

    # Try each possible fixed code
    for fixed_code in possible_codes:
        if fixed_code in subject_dict:
            return subject_dict[fixed_code]

    # Try medium format fallback (remove first 2 digits)
    medium_code = code[2:]  # Remove first 2 digits
    result = lookup_medium_format(medium_code, subject_dict)
    if result:
        return result

    # Try short format fallback (extract last part)
    match = re.match(r"^\d{2}[A-Z]\d{1,2}([A-Z]{2}\d{3,4})$", code)
    if match:
        short_code = match.group(1)
        return lookup_short_format(short_code, subject_dict)

    return None


def lookup_medium_format(code, subject_dict):
    """Lookup medium format with fallback to short format."""
    # Direct lookup first
    if code in subject_dict:
        return subject_dict[code]

    # Try short format fallback (extract last part)
    match = re.match(r"^[A-Z]\d{1,2}([A-Z]{2}\d{3,4})$", code)
    if match:
        short_code = match.group(1)
        return lookup_short_format(short_code, subject_dict)

    return None


def lookup_short_format(code, subject_dict):
    """Lookup short format - direct lookup only."""
    # Fix 4 digits to 3 if needed
    match = re.match(r"^([A-Z]{2})(\d{3,4})$", code)
    if match:
        chars = match.group(1)
        digits = match.group(2)
        if len(digits) == 4:
            digits = digits[:3]
        fixed_code = chars + digits
        return subject_dict.get(fixed_code)

    return subject_dict.get(code)
