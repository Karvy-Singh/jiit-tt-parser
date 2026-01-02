import json

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from jiit_tt_parser.utils.utils import max_bounds

PATH = "faculty.xlsx"


def search_bounds(sheet: Worksheet):
    row, col = max_bounds(sheet)

    for i in range(1, row + 1):
        for j in range(1, col + 1):
            if "Time Table Team" in str(sheet.cell(row=i, column=j).value):
                return row, j - 1

    return row, col


def generate_faculty_map(path: str):
    faculty_map = {}

    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    if sheet is None:
        return faculty_map

    row, col = search_bounds(sheet)

    for j in range(1, col + 1, 2):
        for i in range(1, row + 1):
            if ((code := sheet.cell(i, j).value) is not None) and (
                (name := sheet.cell(i, j + 1).value) is not None
            ):
                faculty_map.update({str(code): str(name)})

    return faculty_map


def parse_down(sheet: Worksheet, curr: int, curc: int):
    faculty_map = {}
    while (v := sheet.cell(curr, curc).value) is not None:
        v = str(v)
        split_word = ":"
        if "-" in v:
            split_word = "-"
        elif ";" in v:  # there is a typo in the spreadsheet
            split_word = ";"

        code, fac = v.split(split_word)
        faculty_map.update({code.strip(): fac.strip()})
        curr += 1

    return faculty_map


# def generate_faculty_128(sheet: Worksheet, row: int, col: int):
#     faculty_map = {}
#     for i in range(1, row+1):
#         for j in range(1, col+1):
#             v = sheet.cell(i, j).value
#
#             if (str(v).strip().startswith("Faculty Abbreviation")):
#                 faculty_map.update(parse_down_bca_N_128(sheet, i+1, j))
#
#     return faculty_map


def get_faculty_map_from_sem1(path):
    faculty_map = {}
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    if sheet is None:
        return faculty_map

    row, col = max_bounds(sheet)
    for i in range(1, row + 1):
        for j in range(1, col + 1):
            v = sheet.cell(i, j).value

            if str(v) == "Faculty Abbreviation with Names":
                faculty_map.update(parse_down(sheet, i + 1, j))

    return faculty_map


def parse_down_bca_N_128(sheet, r, c):
    faculty_map = {}
    while (v := sheet.cell(r, c).value) is not None:
        v2 = sheet.cell(r, c + 1).value
        faculty_map.update({str(v).strip(): str(v2).strip()})
        r += 1

    return faculty_map


def parse_down_128_sem4(sheet, r, c):
    faculty_map = {}
    while (v := sheet.cell(r, c).value) is not None:
        v2 = sheet.cell(r, c + 1).value
        faculty_map.update({str(v2).strip(): str(v).strip()})
        r += 1

    return faculty_map


def generate_faculty_map_from_bca1_N_128(sheet: Worksheet, row: int, col: int):
    faculty_map = {}
    for i in range(1, row + 1):
        for j in range(1, col + 1):
            v = sheet.cell(i, j).value

            if str(v).strip().startswith("Faculty Abbreviation"):
                faculty_map.update(parse_down_bca_N_128(sheet, i + 1, j))

    return faculty_map


def generate_faculty_map_from_128_sem4(sheet: Worksheet, row: int, col: int):
    faculty_map = {}
    for i in range(1, row + 1):
        for j in range(1, col + 1):
            v = sheet.cell(i, j).value

            if str(v).strip().startswith("Faculty Names"):
                faculty_map.update(parse_down_128_sem4(sheet, i + 1, j))

    return faculty_map


def get_faculty_map_from_bca1_N_128(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    if sheet is None:
        return {}

    r, c = max_bounds(sheet)
    faculty_map = generate_faculty_map_from_bca1_N_128(sheet, r, c)

    return faculty_map


def get_faculty_map_from_128_sem4(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    if sheet is None:
        return {}

    r, c = max_bounds(sheet)
    faculty_map = generate_faculty_map_from_128_sem4(sheet, r, c)

    return faculty_map


def get_faculty_map(
    fac1_xl_path: str,
    fac2_xl_path: str,
    sem1_xl_path: str,
    bca1_xl_path: str,
    fac128_xl_path: str,
):
    faculty_map = {}

    faculty_map.update(generate_faculty_map(fac1_xl_path))
    faculty_map.update(get_faculty_map_from_sem1(sem1_xl_path))
    faculty_map.update(generate_faculty_map(fac2_xl_path))
    faculty_map.update(get_faculty_map_from_bca1_N_128(bca1_xl_path))
    faculty_map.update(get_faculty_map_from_bca1_N_128(fac128_xl_path))

    return faculty_map
