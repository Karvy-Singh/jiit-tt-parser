import xls2xlsx 
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import openpyxl
import requests


def cvt_xls_to_xlsx(src_file_path, dst_file_path) -> None:
    xls2xlsx.XLS2XLSX(src_file_path).to_xlsx(dst_file_path)

def max_bounds(sheet: Worksheet) -> tuple[int, int]:
    row, col = sheet.max_row, sheet.max_column
    r, c = 1, 1
    for i in range(1, row+1):
        for j in range(1, col+1):
            value = sheet.cell(row=i, column=j).value
            if value is not None:
                if i > r:
                    r = i
                if j > c:
                    c = j
    return r,c

def print_worksheet(sheet: Worksheet, row: int, column: int):
    print("|", end="")
    for i in range(1, row+1):
        for j in range(1, column+1):
            value = sheet.cell(row=i, column=j).value
            print(value, "|", end="")
        print("\n|", end="")

    print("|", end="")

def is_empty_row(sheet: Worksheet, row: int, cols: int):
    for i in range(1, cols+1):
        if sheet.cell(row, i).value is not None:
            return False

    return True


def download(url: str, save_as: str, block_size: int = 1024*10):
    r = requests.get(url, stream=True)
    size = r.headers.get("Content-Length") or r.headers.get("content-length")
    
    with open(save_as, "wb+") as f:
        for data in r.iter_content(block_size):
            f.write(data)
    
    return size

def load_worksheet(path: str) -> tuple[Worksheet, int, int] | None:
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    if sheet is None:
        return None

    r, c = max_bounds(sheet)
    return sheet, r, c


def are_cells_in_same_merged_group(worksheet: Worksheet, cell1: Cell, cell2: Cell):
    """
    Check if two openpyxl cell objects are part of the same merged cells group.
    
    Args:
        worksheet: openpyxl worksheet object
        cell1: openpyxl Cell object
        cell2: openpyxl Cell object
    
    Returns:
        bool: True if cells are in the same merged group, False otherwise
    """
    
    # Get row and column from cell objects
    row1, col1 = cell1.row, cell1.column
    row2, col2 = cell2.row, cell2.column
    
    # Check all merged cell ranges in the worksheet
    for merged_range in worksheet.merged_cells.ranges:
        # Check if both cells fall within this merged range
        if (merged_range.min_row <= row1 <= merged_range.max_row and
            merged_range.min_col <= col1 <= merged_range.max_col and
            merged_range.min_row <= row2 <= merged_range.max_row and
            merged_range.min_col <= col2 <= merged_range.max_col):
            return True
    
    return False

# Helper function to find which merged group a cell belongs to (if any)
def get_merged_range_for_cell(worksheet: Worksheet, cell: Cell):
    """
    Find the merged range that contains the given cell object.
    
    Args:
        worksheet: openpyxl worksheet object
        cell: openpyxl Cell object
    
    Returns:
        MergedCellRange object if cell is in a merged range, None otherwise
    """
    row, col = cell.row, cell.column
    
    for merged_range in worksheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return merged_range
    
    return None
