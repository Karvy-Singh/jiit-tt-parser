import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from cache import get_cache_file
import json
from utils import max_bounds


PATH = "faculty.xlsx"


def search_bounds(sheet: Worksheet):
    row, col = max_bounds(sheet)

    for i in range(1, row+1):
        for j in range(1, col+1):
            if "Time Table Team" in str(sheet.cell(row=i, column=j).value):
                return row, j-1

    return row, col

def generate_faculty_map(sheet: Worksheet, row: int, col: int):
    faculty_map = {}

    for j in range(1, col+1, 2):
        for i in range(1, row+1):
            if ((code:=sheet.cell(i, j).value) is not None) and ((name:=sheet.cell(i, j+1).value) is not None):
                faculty_map.update({str(code): str(name)})

    return faculty_map

# def faculty_map(def faculty_map())


def cache_faculty_map(map: dict):
    FAC_PATH = get_cache_file("faculty.json")
    with open(FAC_PATH, "w+") as f:
        json.dump(map, f)


def parse_down(sheet: Worksheet, curr: int, curc: int):
    faculty_map = {}
    while ((v:=sheet.cell(curr, curc).value) is not None):
        v = str(v)
        split_word = ":"
        if "-" in v:
            split_word = "-"
        elif ";" in v: # there is a typo in the spreadsheet
            split_word = ";"
        
        code, fac = v.split(split_word)
        faculty_map.update({code.strip(): fac.strip()})
        curr += 1
    
    return faculty_map

def generate_faculty_map_from_sem1(sheet: Worksheet, row: int, col: int):
    faculty_map = {}
    for i in range(1, row+1):
        for j in range(1, col+1):
            v = sheet.cell(i, j).value

            if (str(v) == "Faculty Abbreviation with Names"):
                faculty_map.update(parse_down(sheet, i+1, j))

    return faculty_map


def get_faculty_map_from_sem1(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    r, c = max_bounds(sheet)
    faculty_map = generate_faculty_map_from_sem1(sheet, r, c)

    return faculty_map

def get_faculty_map(fac1_xl_path: str, fac2_xl_path: str, sem1_xl_path: str):
    wb = openpyxl.load_workbook(fac1_xl_path)
    wb2 = openpyxl.load_workbook(fac2_xl_path)

    sheet = wb.active
    sheet2 = wb2.active

    r, c = search_bounds(sheet)
    r2, c2 = search_bounds(sheet2)

    faculty_map = generate_faculty_map(sheet, r, c)
    faculty_map.update(get_faculty_map_from_sem1(sem1_xl_path))
    faculty_map.update(generate_faculty_map(sheet2, r2, c2))


    return faculty_map

if __name__ == "__main__":
    f = get_faculty_map(PATH)

    from pprint import  pprint
    pprint(f)
