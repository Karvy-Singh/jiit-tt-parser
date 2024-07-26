import xls2xlsx 
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl
import requests

def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    xls2xlsx.XLS2XLSX(src_file_path).to_xlsx(dst_file_path)

def max_bounds(sheet: Worksheet):
    row, col = sheet.max_row, sheet.max_column
    r, c = 1, 1
    for i in range(1, row+1):
        for j in range(1, col+1):
            value = sheet.cell(row=i, column=j).value
            if value is not None:
                if i > r:
                    r = i
                if j > c:
                    c = j
    return r,c

def print_worksheet(sheet: Worksheet, row, column):
    print("|", end="")
    for i in range(1, row+1):
        for j in range(1, column+1):
            value = sheet.cell(row=i, column=j).value
            print(value, "|", end="")
        print("\n|", end="")

    print("|", end="")

def is_empty_row(sheet: Worksheet, row, cols):
    for i in range(1, cols+1):
        if sheet.cell(row, i).value is not None:
            return False

    return True


def download(url: str, save_as: str, block_size: int = 1024*10):
    r = requests.get(url, stream=True)
    size = r.headers["Content-Length"]
    
    with open(save_as, "wb") as f:
        for data in r.iter_content(block_size):
            f.write(data)
    
    return size
;
def load_worksheet(path: str):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    r, c = max_bounds(sheet)
    return sheet, r, c
